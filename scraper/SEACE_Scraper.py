"""
SEACE Scraper - Version con ChromeDriver directo (sin webdriver-manager)
Usa el ChromeDriver que ya tienes instalado en el sistema.
"""

import time, sys, os, csv, json, re
from datetime import datetime


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    USE_PANDAS = True
except ImportError:
    USE_PANDAS = False

URL     = "https://prod2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml"
ANO     = "2026"
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
TS      = datetime.now().strftime("%Y%m%d_%H%M%S")

def log(msg, nivel="INFO"):
    icons = {"INFO":"[·]", "OK":"[OK]", "ERR":"[!!]", "WARN":"[!]"}
    print(f"  {icons.get(nivel,'[ ]')} {msg}", flush=True)

def encontrar_chrome():
    """Busca Chrome/Edge en rutas comunes de Windows."""
    rutas_chrome = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    rutas_edge = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    ]

    for ruta in rutas_chrome:
        if os.path.exists(ruta):
            return ruta, "chrome"

    for ruta in rutas_edge:
        if os.path.exists(ruta):
            return ruta, "edge"

    return None, None

def encontrar_chromedriver():
    """Busca chromedriver.exe en el sistema."""
    # 1. En la misma carpeta que el script
    local = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver.exe")
    if os.path.exists(local):
        return local

    # 2. Cache de webdriver-manager
    import glob
    wdm = os.path.expanduser("~/.wdm")
    drivers = glob.glob(wdm + "/**/chromedriver.exe", recursive=True)
    if drivers:
        # El mas reciente
        drivers.sort(key=os.path.getmtime, reverse=True)
        return drivers[0]

    # 3. En el PATH
    import shutil
    en_path = shutil.which("chromedriver")
    if en_path:
        return en_path

    return None

def crear_driver_automatico():
    """Intenta crear el driver de forma automatica."""
    log("Configurando navegador...")

    chrome_bin, _ = encontrar_chrome()
    driver_path   = encontrar_chromedriver()

    log(f"Chrome : {chrome_bin or 'No encontrado'}")
    log(f"Driver : {driver_path or 'No encontrado'}")

    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--lang=es-PE")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    if chrome_bin:
        opts.binary_location = chrome_bin

    # Estrategia 1: driver directo (mas confiable - sin webdriver-manager)
    if driver_path:
        try:
            svc = Service(driver_path)
            driver = webdriver.Chrome(service=svc, options=opts)
            log(f"Chrome iniciado con driver directo", "OK")
            return driver
        except Exception as e:
            log(f"Driver directo fallo: {e}", "WARN")

    # Estrategia 2: Sin service (Selenium busca en PATH)
    try:
        driver = webdriver.Chrome(options=opts)
        log("Chrome iniciado (Selenium encontro driver en PATH)", "OK")
        return driver
    except Exception as e:
        log(f"Chrome sin service fallo: {e}", "WARN")

    # Estrategia 3: Edge como fallback
    try:
        from selenium.webdriver.edge.options import Options as EdgeOptions
        from selenium.webdriver.edge.service import Service as EdgeService

        edge_paths = [
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        ]
        edge_opts = EdgeOptions()
        edge_opts.add_argument("--start-maximized")
        edge_opts.add_argument("--disable-blink-features=AutomationControlled")
        edge_opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        edge_opts.add_experimental_option("useAutomationExtension", False)

        for ep in edge_paths:
            if os.path.exists(ep):
                edge_opts.binary_location = ep
                break

        driver = webdriver.Edge(options=edge_opts)
        log("Microsoft Edge iniciado como alternativa", "OK")
        return driver
    except Exception as e:
        log(f"Edge fallo: {e}", "ERR")

    log("No se pudo inicializar ningun navegador", "ERR")
    print()
    print("  SOLUCION:")
    print(f"  Coloca chromedriver.exe en: {os.path.dirname(os.path.abspath(__file__))}")
    print("  Descarga desde: https://googlechromelabs.github.io/chrome-for-testing/")
    sys.exit(1)


def js_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.3)
    try:
        el.click()
    except Exception:
        driver.execute_script("arguments[0].click();", el)


def seleccionar_tab(driver):
    log("Activando tab Procedimientos de Seleccion...")
    intentos = [
        (By.CSS_SELECTOR, "a[href*='tab1']"),
        (By.XPATH, "//a[contains(.,'Procedimientos de Selecci')]"),
        (By.XPATH, "//li[contains(@id,'tab1')]/a"),
        (By.XPATH, "//li[@role='tab'][2]/a"),
    ]
    for by, sel in intentos:
        try:
            el = driver.find_element(by, sel)
            js_click(driver, el)
            time.sleep(3.5)
            log("Tab activado", "OK")
            return True
        except NoSuchElementException:
            continue
        except Exception:
            continue
    log("Tab no hallado, continuando con pagina actual", "WARN")
    return False


def seleccionar_anio_2026(driver):
    log(f"Seleccionando año {ANO}...")
    # Esperar a que el tab AJAX cargue el formulario
    time.sleep(2)

    for intento in range(3):
        selects = driver.find_elements(By.TAG_NAME, "select")
        for s in selects:
            try:
                if not s.is_displayed():
                    continue
                opts = [o.text.strip() for o in s.find_elements(By.TAG_NAME, "option")]
                # Acepta cualquier select que tenga 2026
                if ANO in opts and len(opts) > 3:
                    sid = s.get_attribute("id") or "sin-id"
                    log(f"  Select hallado id='{sid}' | {len(opts)} opciones")
                    sel_obj = Select(s)
                    sel_obj.select_by_visible_text(ANO)
                    # Disparar change event para JSF
                    driver.execute_script(
                        "var e=new Event('change',{bubbles:true}); arguments[0].dispatchEvent(e);", s
                    )
                    time.sleep(2.5)
                    log(f"  Año {ANO} seleccionado", "OK")
                    return True
            except StaleElementReferenceException:
                continue
            except Exception as e:
                log(f"  Error al interactuar con select: {e}", "WARN")

        if intento < 2:
            log(f"  Reintento {intento+1}: esperando que el formulario cargue...", "WARN")
            time.sleep(3)

    log("Dropdown de año no encontrado", "ERR")
    return False


def click_buscar(driver):
    log("Haciendo click en Buscar...")
    intentos_xpath = [
        "//button[contains(.,'Buscar') and not(contains(.,'Avanzada'))]",
        "//button[.//span[text()=' Buscar'] or .//span[text()='Buscar']]",
        "//input[@value='Buscar']",
    ]
    for xpath in intentos_xpath:
        try:
            els = driver.find_elements(By.XPATH, xpath)
            for el in els:
                if el.is_displayed():
                    txt = el.text.strip()
                    log(f"  Boton: '{txt}' | id='{el.get_attribute('id')}'")
                    js_click(driver, el)
                    log("  Click ejecutado, esperando AJAX...", "OK")
                    time.sleep(7)
                    return True
        except Exception:
            continue

    # Fallback: todos los botones visibles
    for btn in driver.find_elements(By.TAG_NAME, "button"):
        try:
            if "buscar" in btn.text.lower() and btn.is_displayed():
                js_click(driver, btn)
                time.sleep(7)
                log("  Boton Buscar clickeado (por texto)", "OK")
                return True
        except Exception:
            continue

    log("Boton Buscar no encontrado", "ERR")
    return False


def cerrar_popup(driver):
    """Cierra cualquier popup/dialogo abierto de SEACE."""
    intentos = [
        (By.CSS_SELECTOR, ".ui-dialog-titlebar-close"),
        (By.CSS_SELECTOR, ".ui-dialog .ui-icon-closethick"),
        (By.XPATH,        "//div[contains(@class,'ui-dialog')]//a[contains(@class,'ui-dialog-titlebar-close')]"),
        (By.XPATH,        "//button[contains(.,'Cerrar') or contains(.,'Close') or contains(.,'OK')]"),
    ]
    for by, sel in intentos:
        try:
            el = driver.find_element(by, sel)
            if el.is_displayed():
                js_click(driver, el)
                time.sleep(0.6)
                return True
        except Exception:
            continue
    # Fallback: tecla Escape
    try:
        from selenium.webdriver.common.keys import Keys
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.4)
    except Exception:
        pass
    return False


def extraer_texto_popup(driver):
    """
    Espera el dialogo de PrimeFaces y extrae el texto de la tabla interna.
    El popup de SEACE tiene estructura:
      .ui-dialog > .ui-dialog-content > .ui-datatable > tbody.ui-datatable-data > tr > td
    Retorna el texto limpio o cadena vacia.
    """
    # 1. Esperar a que el dialogo sea visible
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".ui-dialog"))
        )
    except TimeoutException:
        return ""

    # 2. Esperar a que la tabla interna cargue su contenido (AJAX)
    try:
        WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located(
                (By.CSS_SELECTOR, ".ui-dialog-content .ui-datatable-data td")
            )
        )
    except TimeoutException:
        # Si no hay tabla interna, intentar leer texto directo del contenido
        try:
            el = driver.find_element(By.CSS_SELECTOR, ".ui-dialog-content")
            return el.text.strip()
        except Exception:
            return ""

    # 3. Recoger todos los valores de las celdas de la tabla interna
    try:
        celdas = driver.find_elements(
            By.CSS_SELECTOR, ".ui-dialog-content .ui-datatable-data td"
        )
        valores = [c.text.strip() for c in celdas if c.text.strip()]
        return ", ".join(valores) if valores else ""
    except Exception:
        return ""


def extraer_codigos_popup(driver, tr, cols):
    """
    Para una fila <tr>, detecta los indices de columna SNIP e Inversion,
    hace clic en su icono <a class='ui-commandlink'> y extrae el texto
    del popup (tabla interna .ui-datatable-data td).
    Retorna dict con claves 'Codigo SNIP' y 'Codigo Unico Inversion'.
    """
    resultado = {"Codigo SNIP": "", "Codigo Unico Inversion": ""}

    # Identificar indices de columna por nombre
    idx_snip = next(
        (i for i, c in enumerate(cols) if "snip" in c.lower()), None
    )
    idx_inv = next(
        (i for i, c in enumerate(cols)
         if ("inversion" in c.lower() or "inversion" in c.lower())
         and "codigo" in c.lower()),
        None
    )

    tds = tr.find_elements(By.TAG_NAME, "td")

    for clave, idx in [("Codigo SNIP", idx_snip), ("Codigo Unico Inversion", idx_inv)]:
        if idx is None or idx >= len(tds):
            continue
        td = tds[idx]

        # El icono es un <a class="ui-commandlink"> que contiene un <img>
        icono = None
        for selector in ["a.ui-commandlink", "a[class*='commandlink']", "a", "button"]:
            try:
                ic = td.find_element(By.CSS_SELECTOR, selector)
                if ic.is_displayed():
                    icono = ic
                    break
            except Exception:
                continue

        if not icono:
            log(f"    Sin icono clickeable en columna '{clave}' (idx={idx})", "WARN")
            continue

        try:
            js_click(driver, icono)
            texto = extraer_texto_popup(driver)
            resultado[clave] = texto
            log(f"    {clave}: '{texto}'", "OK" if texto else "WARN")
            cerrar_popup(driver)
            time.sleep(0.8)  # pausa para que el DOM se estabilice
        except Exception as e:
            log(f"    Error popup {clave}: {e}", "WARN")
            cerrar_popup(driver)

    return resultado


def extraer_filas(driver):
    """Extrae filas de la tabla activa, incluyendo popups de Codigo SNIP e Inversion."""
    filas = []
    time.sleep(1.5)
    try:
        # Buscar todas las tablas con contenido
        tablas = driver.find_elements(By.TAG_NAME, "table")
        tabla_max = None
        max_tds = 0
        for t in tablas:
            try:
                tds = t.find_elements(By.TAG_NAME, "td")
                if len(tds) > max_tds:
                    max_tds = len(tds)
                    tabla_max = t
            except Exception:
                continue

        if not tabla_max or max_tds < 3:
            log("No se encontro tabla de resultados", "WARN")
            return filas

        # Encabezados
        ths = tabla_max.find_elements(By.TAG_NAME, "th")
        cols = [th.text.strip() for th in ths if th.text.strip()]
        if not cols:
            cols = [f"Columna_{i+1}" for i in range(10)]

        log(f"  Columnas: {cols}")

        # Determinar si hay columnas SNIP / Inversion para activar modo popup
        tiene_snip = any("snip" in c.lower() for c in cols)
        tiene_inv  = any(
            ("inversion" in c.lower() or "inversión" in c.lower())
            and "codigo" in c.lower()
            for c in cols
        )
        modo_popup = tiene_snip or tiene_inv
        if modo_popup:
            log(f"  Modo popup activado (SNIP={tiene_snip}, Inversion={tiene_inv})")

        # Filas de datos
        trs = tabla_max.find_elements(By.CSS_SELECTOR, "tbody tr")
        if not trs:
            trs = [r for r in tabla_max.find_elements(By.TAG_NAME, "tr")
                   if r.find_elements(By.TAG_NAME, "td")]

        for fila_num, tr in enumerate(trs):
            try:
                tds = tr.find_elements(By.TAG_NAME, "td")
                vals = [td.text.strip() for td in tds]
                if not any(v for v in vals):
                    continue

                # Armar dict base
                row = {}
                for i, v in enumerate(vals):
                    col = cols[i] if i < len(cols) else f"col_{i+1}"
                    row[col] = v

                # Extraer codigos de popups si aplica
                if modo_popup:
                    popups = extraer_codigos_popup(driver, tr, cols)
                    # Sobreescribir las celdas SNIP/Inversion con el texto real del popup
                    for clave_popup, valor_popup in popups.items():
                        if valor_popup:
                            # Buscar columna equivalente en row y reemplazar
                            for col_name in list(row.keys()):
                                if clave_popup == "Codigo SNIP" and "snip" in col_name.lower():
                                    row[col_name] = valor_popup
                                elif clave_popup == "Codigo Unico Inversion" and \
                                     ("inversion" in col_name.lower() or "inversión" in col_name.lower()):
                                    row[col_name] = valor_popup
                            # Agregar tambien como campo extra si el nombre de columna era vacio/icono
                            row[clave_popup] = valor_popup

                # Extraer Ficha de Seleccion:
                # IMPORTANTE: La navegacion invalida todas las referencias DOM actuales.
                # Por eso la Ficha se extrae ULTIMA, despues de todos los popups en pagina.
                log(f"  Extrayendo Ficha de Seleccion fila {fila_num+1}...")
                ficha = extraer_ficha_seleccion(driver, tr)
                if ficha:
                    row["ficha"] = ficha

                filas.append(row)

            except StaleElementReferenceException:
                log(f"  Fila {fila_num+1}: referencia obsoleta, saltando", "WARN")
                continue
            except Exception as e:
                log(f"  Error en fila {fila_num+1}: {e}", "WARN")
                continue

    except Exception as e:
        log(f"Error extrayendo tabla: {e}", "WARN")

    return filas


def _texto(el):
    """Devuelve el texto limpio de un elemento o cadena vacia."""
    try:
        return el.text.strip()
    except Exception:
        return ""


def extraer_ficha_seleccion(driver, tr):
    """
    Hace clic en el icono 'Ver Ficha de Seleccion' de la fila <tr>,
    navega a la pagina de detalle y extrae todos los datos estructurados:
      - Campos generales (Convocatoria)
      - Cronograma (etapas con fechas)
      - Documentos por etapa
      - Listado de items
    Regresa a la pagina de resultados al terminar.
    Devuelve dict con todos los datos.
    """
    ficha = {}
    url_original = driver.current_url

    try:
        # ── 1. Encontrar icono Ficha de Seleccion ──
        icono_ficha = None
        for a in tr.find_elements(By.CSS_SELECTOR, "a.ui-commandlink"):
            try:
                titulo = (a.get_attribute("title") or "").lower()
                img_src = ""
                try:
                    img_src = a.find_element(By.TAG_NAME, "img").get_attribute("src") or ""
                except Exception:
                    pass
                if "ficha" in titulo or "sheet_calendar" in img_src:
                    icono_ficha = a
                    break
            except Exception:
                continue

        if not icono_ficha:
            log("    Icono Ficha de Seleccion no encontrado en fila", "WARN")
            return ficha

        log("    Abriendo Ficha de Seleccion...")
        js_click(driver, icono_ficha)
        time.sleep(5)

        if "fichaSeleccion" not in driver.current_url:
            log(f"    URL inesperada: {driver.current_url}", "WARN")
            driver.back()
            time.sleep(3)
            return ficha

        log(f"    Ficha URL: {driver.current_url}", "OK")
        ficha["ficha_url"] = driver.current_url

        # ── 2. Campos generales (pares etiqueta:valor) ──
        try:
            celdas = driver.find_elements(By.CSS_SELECTOR, ".ui-panelgrid-cell")
            i = 0
            while i < len(celdas) - 1:
                clave = _texto(celdas[i]).rstrip(":").strip()
                valor = _texto(celdas[i + 1])
                if clave and valor:
                    ficha[clave] = valor
                i += 2
        except Exception as e:
            log(f"    Campos generales: {e}", "WARN")

        # Fallback: tablas de pares
        if len(ficha) <= 1:
            try:
                for tr_f in driver.find_elements(By.CSS_SELECTOR, ".ui-fieldset-content tr"):
                    tds = tr_f.find_elements(By.TAG_NAME, "td")
                    if len(tds) >= 2:
                        clave = _texto(tds[0]).rstrip(":").strip()
                        valor = _texto(tds[1])
                        if clave and valor:
                            ficha[clave] = valor
            except Exception:
                pass

        # ── 3. Cronograma ──
        cronograma = []
        try:
            tablas_crono = driver.find_elements(
                By.CSS_SELECTOR,
                "[id*='ronograma'] table, [id*='Cronograma'] table"
            )
            for tabla in tablas_crono:
                ths = [_texto(th) for th in tabla.find_elements(By.TAG_NAME, "th") if _texto(th)]
                for tr_c in tabla.find_elements(By.CSS_SELECTOR, "tbody tr"):
                    tds = [_texto(td) for td in tr_c.find_elements(By.TAG_NAME, "td")]
                    if any(tds):
                        cronograma.append(dict(zip(ths, tds)) if ths and len(ths)==len(tds) else tds)
        except Exception as e:
            log(f"    Cronograma: {e}", "WARN")
        if cronograma:
            ficha["cronograma"] = cronograma

        # ── 4. Documentos ──
        documentos = []
        try:
            tablas_doc = driver.find_elements(
                By.CSS_SELECTOR,
                "[id*='ocumento'] table, [id*='Documento'] table"
            )
            for tabla in tablas_doc:
                ths = [_texto(th) for th in tabla.find_elements(By.TAG_NAME, "th") if _texto(th)]
                for tr_d in tabla.find_elements(By.CSS_SELECTOR, "tbody tr"):
                    tds = tr_d.find_elements(By.TAG_NAME, "td")
                    fila_doc = {}
                    for i, td in enumerate(tds):
                        col = ths[i] if i < len(ths) else f"col_{i+1}"
                        try:
                            href = td.find_element(By.TAG_NAME, "a").get_attribute("href")
                            fila_doc[col] = href or _texto(td)
                        except Exception:
                            fila_doc[col] = _texto(td)
                    if any(fila_doc.values()):
                        documentos.append(fila_doc)
        except Exception as e:
            log(f"    Documentos: {e}", "WARN")
        if documentos:
            ficha["documentos"] = documentos

        # ── 5. Items (expandir panel si esta colapsado) ──
        items = []
        try:
            for btn in driver.find_elements(By.CSS_SELECTOR, "a, .ui-fieldset-toggler"):
                try:
                    if "item" in _texto(btn).lower():
                        js_click(driver, btn)
                        time.sleep(1.5)
                        break
                except Exception:
                    pass
            tablas_item = driver.find_elements(
                By.CSS_SELECTOR, "[id*='tem'] table"
            )
            for tabla in tablas_item:
                ths = [_texto(th) for th in tabla.find_elements(By.TAG_NAME, "th") if _texto(th)]
                for tr_i in tabla.find_elements(By.CSS_SELECTOR, "tbody tr"):
                    tds = [_texto(td) for td in tr_i.find_elements(By.TAG_NAME, "td")]
                    if any(tds):
                        items.append(dict(zip(ths, tds)) if ths and len(ths)==len(tds) else tds)
        except Exception as e:
            log(f"    Items: {e}", "WARN")
        if items:
            ficha["items"] = items

        log(
            f"    Ficha OK: {len(ficha)} campos | "
            f"{len(cronograma)} etapas | {len(documentos)} docs | {len(items)} items",
            "OK"
        )

    except Exception as e:
        log(f"    Error extrayendo Ficha: {e}", "ERR")

    finally:
        # ── 6. Regresar a resultados ──
        try:
            btn = driver.find_element(
                By.XPATH,
                "//a[normalize-space(.)='Regresar'] | //button[normalize-space(.)='Regresar']"
            )
            js_click(driver, btn)
            time.sleep(4)
        except Exception:
            driver.get(url_original)
            time.sleep(5)

    return ficha


def siguiente_pagina(driver):
    """Click en boton 'Siguiente' del paginador."""
    sels = [
        ".ui-paginator-next:not(.ui-state-disabled)",
        "a.ui-paginator-next",
        "//a[contains(@class,'paginator-next') and not(contains(@class,'disabled'))]",
        "//a[contains(@title,'Siguiente')]",
    ]
    for sel in sels:
        try:
            by = By.XPATH if sel.startswith("//") else By.CSS_SELECTOR
            el = driver.find_element(by, sel)
            classes = el.get_attribute("class") or ""
            if "disabled" in classes or "ui-state-disabled" in classes:
                return False
            js_click(driver, el)
            time.sleep(3.5)
            return True
        except NoSuchElementException:
            continue
    return False


def screenshot(driver, nombre):
    try:
        ruta = os.path.join(DESKTOP, f"SEACE_{nombre}.png")
        driver.save_screenshot(ruta)
    except Exception:
        pass


def guardar(datos):
    if not datos:
        log("Sin datos para guardar", "WARN")
        return

    base = os.path.join(DESKTOP, f"SEACE_Procedimientos_{ANO}_{TS}")

    # CSV
    cols = list(datos[0].keys())
    with open(base + ".csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader(); w.writerows(datos)
    log(f"CSV: {base}.csv", "OK")

    # JSON
    with open(base + ".json", "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
    log(f"JSON: {base}.json", "OK")

    # Excel
    if USE_PANDAS:
        df = pd.DataFrame(datos)
        with pd.ExcelWriter(base + ".xlsx", engine="openpyxl") as ew:
            df.to_excel(ew, index=False, sheet_name="SEACE_2026")
            ws = ew.sheets["SEACE_2026"]
            fill = PatternFill("solid", fgColor="1F4E79")
            for i, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=i)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                max_l = max(len(str(col)), df[col].astype(str).str.len().max() if len(df)>0 else 10)
                ws.column_dimensions[get_column_letter(i)].width = min(max_l + 3, 55)
        log(f"Excel: {base}.xlsx", "OK")

    # HTML interactivo
    ruta_html = base + ".html"
    generar_html(datos, cols, ruta_html)

    # Preview consola
    print()
    print("  " + "─"*60)
    print(f"  COLUMNAS : {list(datos[0].keys())}")
    print(f"  REGISTROS: {len(datos)}")
    print("  " + "─"*60)
    print("  PREVIEW - primeros 3 registros:")
    for i, row in enumerate(datos[:3], 1):
        print(f"\n  [{i}]")
        for k, v in row.items():
            if v: print(f"       {k}: {v}")


def generar_html(datos, cols, ruta_html):
    import html as html_mod
    
    total = len(datos)
    col_objeto = next((c for c in cols if "objeto" in c.lower() and "contrat" in c.lower()), None)
    col_entidad = next((c for c in cols if "entidad" in c.lower() or "sigla" in c.lower()), None)
    col_nomenclatura = next((c for c in cols if "nomenclatura" in c.lower() or "nomencla" in c.lower()), None)

    tipo_count = {}
    if col_objeto:
        for row in datos:
            t = row.get(col_objeto, "Otro").strip() or "No especificado"
            tipo_count[t] = tipo_count.get(t, 0) + 1

    entidades_unicas = len(set(r.get(col_entidad, "") for r in datos if col_entidad)) if col_entidad else 0

    tipo_colores = {
        "Bien":              ("#0ea5e9", "#e0f2fe"),
        "Obra":              ("#f59e0b", "#fef3c7"),
        "Servicio":          ("#10b981", "#d1fae5"),
        "Consultoría de Obra": ("#8b5cf6", "#ede9fe"),
        "No especificado":   ("#64748b", "#f1f5f9"),
    }

    filas_html = []
    for idx, row in enumerate(datos):
        tipo_val = row.get(col_objeto, "").strip() if col_objeto else ""
        color_txt, color_bg = tipo_colores.get(tipo_val, ("#64748b", "#f8fafc"))

        celdas = []
        for c in cols:
            val = html_mod.escape(str(row.get(c, "") or ""))
            if c == col_objeto and val:
                celdas.append(f'<td><span class="badge" style="background:{color_bg};color:{color_txt};border:1px solid {color_txt}22">{val}</span></td>')
            elif c == col_nomenclatura and val:
                celdas.append(f'<td><code class="nomen">{val}</code></td>')
            elif c and "fecha" in c.lower() and val:
                celdas.append(f'<td><span class="fecha">{val}</span></td>')
            else:
                celdas.append(f'<td>{val}</td>')

        clase = "odd" if idx % 2 == 0 else "even"
        filas_html.append(f'<tr class="{clase}" data-idx="{idx}">{"".join(celdas)}</tr>')
    filas_str = "\n".join(filas_html)

    headers_html = "".join(f'<th onclick="sortTable({i})" title="Ordenar por esta columna">{html_mod.escape(c)} <span class="sort-icon">⇅</span></th>' for i, c in enumerate(cols))

    stats_cards = f"""
        <div class="stat-card">
          <div class="stat-number">{total:,}</div>
          <div class="stat-label">Total Registros</div>
        </div>
        <div class="stat-card">
          <div class="stat-number">{entidades_unicas:,}</div>
          <div class="stat-label">Entidades Únicas</div>
        </div>
    """
    for tipo, cnt in sorted(tipo_count.items(), key=lambda x: -x[1])[:4]:
        ct, cb = tipo_colores.get(tipo, ("#64748b", "#f1f5f9"))
        stats_cards += f"""
        <div class="stat-card" style="border-top:4px solid {ct}">
          <div class="stat-number" style="color:{ct}">{cnt:,}</div>
          <div class="stat-label">{html_mod.escape(tipo)}</div>
        </div>"""

    opciones_tipo = '<option value="">Todos los tipos</option>'
    for tipo in sorted(tipo_count.keys()):
        opciones_tipo += f'<option value="{html_mod.escape(tipo)}">{html_mod.escape(tipo)} ({tipo_count[tipo]})</option>'

    col_objeto_idx = cols.index(col_objeto) if col_objeto and col_objeto in cols else -1
    col_entidad_idx = cols.index(col_entidad) if col_entidad and col_entidad in cols else -1

    datos_json = json.dumps(datos, ensure_ascii=False)

    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>SEACE — Procedimientos de Selección {ANO}</title>
  <style>
    :root {{
      --bg:          #f4f6f9;
      --surface:     #ffffff;
      --surface2:    #f8fafc;
      --border:      #cbd5e1;
      --accent:      #0284c7; /* Azul SEACE */
      --accent-orange: #f59e0b; /* Naranja SEACE */
      --text:        #1f2937;
      --text-muted:  #4b5563;
      --odd:         #ffffff;
      --even:        #f8fafc;
      --hover:       #e0f2fe;
      --radius:      4px;
    }}
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; }}
    header {{ background: #fff; padding: 12px 30px; border-top: 5px solid var(--accent-orange); border-bottom: 1px solid var(--border); position: sticky; top: 0; z-index: 100; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
    .header-top {{ display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 12px; width: 100%; }}
    .logo-area {{ display: flex; align-items: center; gap: 12px; }}
    .logo-icon {{ width: 36px; height: 36px; background: #e0f2fe; color: var(--accent); border: 1px solid #bae6fd; display: flex; align-items: center; justify-content: center; font-size: 18px; border-radius: var(--radius); }}
    h1 {{ font-size: 1.25rem; font-weight: 600; color: var(--accent); line-height: 1.2; margin: 0; }}
    h1 span {{ display: block; font-size: 0.72rem; font-weight: 400; color: var(--text-muted); margin-top: 2px; }}
    .live-badge {{ display: inline-flex; align-items: center; background: #dcfce7; border: 1px solid #86efac; color: #166534; padding: 3px 8px; border-radius: 4px; font-size: 0.65rem; font-weight: 600; }}
    
    .stats-section {{ padding: 15px 30px; background: var(--surface); border-bottom: 1px solid var(--border); }}
    .stats-grid {{ display: flex; gap: 12px; flex-wrap: wrap; }}
    .stat-card {{ background: #fff; border: 1px solid var(--border); border-left: 4px solid var(--accent); border-radius: 4px; padding: 10px 14px; min-width: 130px; flex: 1; max-width: 180px; box-shadow: 0 1px 2px rgba(0,0,0,0.02); }}
    .stat-number {{ font-size: 1.35rem; font-weight: 700; color: var(--text); }}
    .stat-label {{ font-size: 0.65rem; color: var(--text-muted); text-transform: uppercase; margin-top: 2px; }}
    
    .controls-section {{ padding: 12px 30px; background: var(--bg); display: flex; gap: 10px; flex-wrap: wrap; align-items: center; border-bottom: 1px solid var(--border); }}
    .search-box {{ position: relative; flex: 1; min-width: 240px; }}
    .search-box input {{ width: 100%; padding: 7px 10px 7px 32px; border: 1px solid var(--border); border-radius: var(--radius); font-size: 0.8rem; outline: none; }}
    .search-box input:focus {{ border-color: var(--accent); box-shadow: 0 0 0 2px rgba(2,132,199,0.2); }}
    .search-box::before {{ content: '🔍'; position: absolute; left: 10px; top: 50%; transform: translateY(-50%); font-size: 0.75rem; color: #9ca3af; }}
    select.filter-select {{ padding: 7px 10px; border: 1px solid var(--border); border-radius: var(--radius); font-size: 0.8rem; outline: none; background: #fff; min-width: 160px; }}
    .btn {{ padding: 7px 14px; border: none; border-radius: var(--radius); font-size: 0.8rem; cursor: pointer; font-weight: 500; transition: background 0.15s; }}
    .btn-primary {{ background: var(--accent); color: #fff; }}
    .btn-primary:hover {{ background: #0369a1; }}
    .btn-secondary {{ background: #fff; border: 1px solid var(--border); color: var(--text); }}
    .btn-secondary:hover {{ background: #f1f5f9; }}
    .result-count {{ font-size: 0.75rem; color: var(--text-muted); margin-left: auto; }}
    
    .table-container {{ padding: 20px 30px; overflow-x: auto; background: var(--surface); }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.7rem; border: 1px solid var(--border); min-width: 1200px; }}
    thead th {{ background: #f8fafc; color: var(--accent); padding: 8px 6px; text-align: center; font-weight: 600; border: 1px solid var(--border); position: sticky; top: 65px; z-index: 10; cursor: pointer; user-select: none; }}
    thead th:hover {{ background: #f1f5f9; }}
    .sort-icon {{ font-size: 0.6rem; color: #94a3b8; font-weight: normal; margin-left: 2px; }}
    tbody tr.odd {{ background: var(--odd); }}
    tbody tr.even {{ background: var(--even); }}
    tbody tr:hover {{ background: var(--hover); }}
    td {{ padding: 6px 8px; border: 1px solid var(--border); vertical-align: middle; line-height: 1.3; text-align: center; color: #334155; }}
    /* Let's left-align the entity name and the description */
    td:nth-child(2), td:nth-child(7) {{ text-align: left; }}
    
    .badge {{ display: inline-block; padding: 2px 6px; border-radius: 3px; font-size: 0.65rem; font-weight: 600; background: #f1f5f9; border: 1px solid #cbd5e1; color: var(--text-muted); }}
    .nomen {{ font-family: monospace; font-size: 0.7rem; color: var(--text); background: #f8fafc; padding: 2px 4px; border: 1px solid #e2e8f0; border-radius: 3px; }}
    .fecha {{ color: #4b5563; }}
    .empty-state {{ text-align: center; padding: 40px; color: var(--text-muted); display: none; }}
    
    .modal-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.5); z-index: 1000; align-items: center; justify-content: center; }}
    .modal-overlay.open {{ display: flex; }}
    .modal {{ background: #fff; border-radius: 6px; max-width: 600px; width: 90%; max-height: 85vh; overflow-y: auto; box-shadow: 0 10px 25px rgba(0,0,0,0.15); border: 1px solid var(--border); }}
    .modal-header {{ padding: 14px 20px; border-bottom: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; position: sticky; top: 0; background: #fff; z-index: 1; }}
    .modal-close {{ background: none; border: none; font-size: 1.4rem; cursor: pointer; color: #94a3b8; }}
    .modal-close:hover {{ color: #1f2937; }}
    .modal-body {{ padding: 16px 20px; }}
    .modal-field {{ display: grid; grid-template-columns: 140px 1fr; gap: 12px; padding: 8px 0; border-bottom: 1px solid #f1f5f9; align-items: start; text-align: left; }}
    .field-label {{ font-size: 0.7rem; color: var(--accent); font-weight: 600; text-transform: uppercase; }}
    .field-value {{ font-size: 0.75rem; color: var(--text); line-height: 1.4; }}
    footer {{ text-align: center; padding: 12px; font-size: 0.7rem; color: var(--text-muted); background: var(--bg); border-top: 1px solid var(--border); }}
    footer a {{ color: var(--accent); text-decoration: none; }}
  </style>
</head>
<body>

<header>
  <div class="header-top">
    <div class="logo-area">
      <div class="logo-icon">🏛️</div>
      <h1>SEACE — Procedimientos de Selección
        <span>Año de Convocatoria {ANO} · Generado el {(datetime.now().strftime('%d/%m/%Y a las %H:%M'))}</span>
      </h1>
    </div>
    <div class="header-meta">
      <div class="live-badge">DATOS EXTRAÍDOS</div>
      <div style="margin-top:6px">Fuente: prod2.seace.gob.pe</div>
    </div>
  </div>
</header>

<section class="stats-section">
  <div class="stats-grid">
    {stats_cards}
  </div>
</section>

<div class="controls-section">
  <div class="search-box">
    <input type="text" id="searchInput" placeholder="Buscar en todos los campos..." oninput="filtrar()" autocomplete="off">
  </div>
  <select class="filter-select" id="tipoFilter" onchange="filtrar()">
    {opciones_tipo}
  </select>
  <button class="btn btn-secondary" onclick="limpiarFiltros()">✕ Limpiar</button>
  <button class="btn btn-primary" onclick="exportarCSV()">⬇ Exportar CSV</button>
  <div class="result-count">Mostrando <span id="countVisible">{total}</span> de {total} registros</div>
</div>

<div class="table-container">
  <table id="mainTable">
    <thead>
      <tr>
        <th onclick="sortTable(0)" title="N°"># <span class="sort-icon"></span></th>
        {headers_html}
        <th style="cursor:default">Ver</th>
      </tr>
    </thead>
    <tbody id="tableBody">
      {filas_str}
    </tbody>
  </table>
  <div class="empty-state" id="emptyState">
    <div class="icon">🔍</div>
    <p>No se encontraron resultados para tu búsqueda.</p>
  </div>
</div>

<div class="modal-overlay" id="modalOverlay" onclick="cerrarModal(event)">
  <div class="modal" id="modal">
    <div class="modal-header">
      <h3 id="modalTitle">Detalle del procedimiento</h3>
      <button class="modal-close" onclick="document.getElementById('modalOverlay').classList.remove('open')">×</button>
    </div>
    <div class="modal-body" id="modalBody"></div>
  </div>
</div>

<footer>
  Reporte generado automáticamente por <strong>SEACE Scraper</strong> ·
  Datos de <a href="https://prod2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml" target="_blank">seace.gob.pe</a> ·
  {(datetime.now().strftime('%d/%m/%Y %H:%M'))}
</footer>

<script>
  const DATOS = {datos_json};
  const COLS  = {json.dumps(cols, ensure_ascii=False)};
  const COL_OBJETO_IDX  = {col_objeto_idx};
  const COL_ENTIDAD_IDX = {col_entidad_idx};

  let sortCol = -1, sortAsc = true;

  document.querySelectorAll('#tableBody tr').forEach((tr, i) => {{
    const td = document.createElement('td');
    td.innerHTML = `<button onclick="abrirModal(${{i}})" style="padding:4px 10px;border:1px solid #334155;background:#1e293b;color:#94a3b8;border-radius:6px;cursor:pointer;font-size:0.75rem" title="Ver detalle">👁</button>`;
    tr.appendChild(td);
  }});

  function filtrar() {{
    const q     = document.getElementById('searchInput').value.toLowerCase().trim();
    const tipo  = document.getElementById('tipoFilter').value;
    const tbody = document.getElementById('tableBody');
    const rows  = Array.from(tbody.querySelectorAll('tr'));
    let vis = 0;

    rows.forEach((tr) => {{
      const txt = tr.textContent.toLowerCase();
      const celdas = tr.querySelectorAll('td');
      const tipoVal = COL_OBJETO_IDX >= 0 && celdas[COL_OBJETO_IDX + 1] ? celdas[COL_OBJETO_IDX + 1].textContent.trim() : '';
      const matchQ    = !q    || txt.includes(q);
      const matchTipo = !tipo || tipoVal === tipo;

      if (matchQ && matchTipo) {{ tr.style.display = ''; vis++; }}
      else                      {{ tr.style.display = 'none'; }}
    }});

    document.getElementById('countVisible').textContent = vis;
    document.getElementById('emptyState').style.display = vis === 0 ? 'block' : 'none';
  }}

  function limpiarFiltros() {{
    document.getElementById('searchInput').value = '';
    document.getElementById('tipoFilter').value  = '';
    filtrar();
  }}

  function sortTable(col) {{
    if (sortCol === col) sortAsc = !sortAsc;
    else {{ sortCol = col; sortAsc = true; }}

    document.querySelectorAll('thead th').forEach((th, i) => {{
      th.classList.remove('asc', 'desc');
      if (i === col + 1) th.classList.add(sortAsc ? 'asc' : 'desc');
    }});

    const tbody = document.getElementById('tableBody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    rows.sort((a, b) => {{
      const tds_a = a.querySelectorAll('td');
      const tds_b = b.querySelectorAll('td');
      const va = tds_a[col] ? tds_a[col].textContent.trim() : '';
      const vb = tds_b[col] ? tds_b[col].textContent.trim() : '';
      const na = parseFloat(va.replace(/[^0-9.\\-]/g,'')), nb = parseFloat(vb.replace(/[^0-9.\\-]/g,''));
      if (!isNaN(na) && !isNaN(nb)) return sortAsc ? na - nb : nb - na;
      return sortAsc ? va.localeCompare(vb,'es') : vb.localeCompare(va,'es');
    }});
    rows.forEach(r => tbody.appendChild(r));
    let vis = 0;
    rows.forEach(r => {{ if (r.style.display !== 'none') {{ r.className = vis % 2 === 0 ? 'odd' : 'even'; vis++; }} }});
    filtrar();
  }}

  function abrirModal(idx) {{
    const row = DATOS[idx];
    if (!row) return;
    const entidad = row[COLS[COL_ENTIDAD_IDX]] || '';
    document.getElementById('modalTitle').textContent = entidad || ('Registro #' + (idx + 1));
    let html = '';
    COLS.forEach(col => {{
      const val = row[col] || '';
      if (!val) return;
      html += `<div class="modal-field"><div class="field-label">${{col}}</div><div class="field-value">${{val}}</div></div>`;
    }});
    document.getElementById('modalBody').innerHTML = html;
    document.getElementById('modalOverlay').classList.add('open');
  }}

  function cerrarModal(e) {{
    if (e.target === document.getElementById('modalOverlay'))
      document.getElementById('modalOverlay').classList.remove('open');
  }}

  document.addEventListener('keydown', e => {{
    if (e.key === 'Escape') document.getElementById('modalOverlay').classList.remove('open');
    if ((e.ctrlKey || e.metaKey) && e.key === 'f') {{ e.preventDefault(); document.getElementById('searchInput').focus(); }}
  }});

  function exportarCSV() {{
    const rows = Array.from(document.querySelectorAll('#tableBody tr')).filter(r => r.style.display !== 'none');
    const header = COLS.join(',');
    const lines = rows.map(tr => {{
      const idx = parseInt(tr.dataset.idx);
      const row = DATOS[idx] || {{}};
      return COLS.map(c => '"' + (row[c] || '').replace(/"/g,'""') + '"').join(',');
    }});
    const blob = new Blob([header + '\\n' + lines.join('\\n')], {{type: 'text/csv;charset=utf-8;'}});
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'SEACE_filtrado_{{ANO}}.csv';
    a.click();
  }}

  document.getElementById('searchInput').addEventListener('focus', function() {{ this.select(); }});
</script>
</body>
</html>"""

    with open(ruta_html, "w", encoding="utf-8") as f:
        f.write(html_content)
    log(f"HTML: {ruta_html}", "OK")

    try:
        import webbrowser
        webbrowser.open(f"file:///{ruta_html.replace(chr(92), '/')}")
        log("Reporte HTML abierto en el navegador", "OK")
    except Exception as e:
        log(f"Abre manualmente: {ruta_html}", "WARN")



# ===== MAIN =====
def main():
    print()
    print("  " + "═"*60)
    print("    SEACE - Buscador Procedimientos de Seleccion")
    print(f"    Año: {ANO} | {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("  " + "═"*60)
    print()

    driver = crear_driver_automatico()
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
    })

    todos      = []
    pagina     = 0
    estado_run = "ERROR"

    try:
        log(f"Abriendo SEACE...")
        driver.get(URL)
        time.sleep(6)
        log(f"Titulo: {driver.title}", "OK")
        screenshot(driver, "01_inicio")

        seleccionar_tab(driver)
        screenshot(driver, "02_tab")

        seleccionar_anio_2026(driver)
        screenshot(driver, "03_anio")

        ok = click_buscar(driver)
        screenshot(driver, "04_busqueda")

        if not ok:
            log("No se hizo click automatico. El Chrome esta abierto.", "WARN")
            log("Por favor haz click en 'Buscar' manualmente y presiona ENTER aqui.", "WARN")
            input("  >>> Presiona ENTER cuando veas los resultados en Chrome...")
            time.sleep(3)

        # Extraer y guardar pagina a pagina
        pagina    = 1
        sin_datos = 0

        while pagina <= 200:
            log(f"Extrayendo pagina {pagina}...")
            filas = extraer_filas(driver)

            if filas:
                todos.extend(filas)
                log(
                    f"  Pag {pagina}: {len(filas)} filas | Acumulado: {len(todos)}",
                    "OK"
                )
                sin_datos = 0
            else:
                sin_datos += 1
                log(f"  Pagina {pagina}: sin filas ({sin_datos}/2)", "WARN")
                if sin_datos >= 2:
                    break

            if not siguiente_pagina(driver):
                log("Fin de paginas.", "OK")
                break
            if pagina >= 1:  # MODO TEST: solo pagina 1
                log("[TEST] Limite de 1 pagina alcanzado. Deteniendo.", "WARN")
                break
            pagina += 1

        estado_run = "OK"

    except KeyboardInterrupt:
        log("Interrumpido por el usuario", "WARN")
        estado_run = "INTERRUMPIDO"
    except Exception as e:
        log(f"Error: {e}", "ERR")
        import traceback; traceback.print_exc()
        screenshot(driver, "ERROR")
        estado_run = "ERROR"
    finally:
        screenshot(driver, "final")
        try:
            driver.quit()
        except Exception:
            pass

    # ── Guardar JSON ──
    if todos:
        ts_local = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_json = os.path.join(DESKTOP, f"SEACE_{ANO}_{ts_local}.json")
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(todos, f, ensure_ascii=False, indent=2)
        log(f"JSON guardado: {ruta_json}", "OK")
    else:
        log("Sin datos para guardar.", "WARN")

    # ── Resumen final ──
    print()
    print("  " + "─"*60)
    print(f"  Registros extraidos : {len(todos)}")
    print(f"  Estado              : {estado_run}")
    if todos:
        print(f"  Archivo JSON        : {ruta_json}")
    print("  " + "─"*60)
    print()

if __name__ == "__main__":
    main()
    input("  Presiona ENTER para salir...")
