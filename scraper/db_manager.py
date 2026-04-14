"""
db_manager.py — Capa de almacenamiento SQLite para SEACE Scraper
================================================================
Gestiona la base de datos seace.db con:
  · Tabla `procedimientos`   — registros scrapeados, deduplicados por
                                (nomenclatura, fecha_publicacion)
  · Tabla `scraping_runs`    — metadatos de cada ejecucion del scraper

Regla de deduplicacion:
  · DUPLICADO REAL   = misma nomenclatura + misma fecha → se ignora
  · RECONVOCATORIA   = misma nomenclatura + fecha distinta → se inserta
"""

import os
import sqlite3
import csv
import json
from datetime import datetime

# ─────────────────────────────────────────────
#  Rutas
# ─────────────────────────────────────────────
DESKTOP    = os.path.join(os.path.expanduser("~"), "Desktop")
DB_DIR     = os.path.join(DESKTOP, "SEACE_DB")
DB_PATH    = os.path.join(DB_DIR, "seace.db")
EXPORT_DIR = os.path.join(DB_DIR, "exports")
LOG_DIR    = os.path.join(DB_DIR, "logs")

for _d in (DB_DIR, EXPORT_DIR, LOG_DIR):
    os.makedirs(_d, exist_ok=True)


# ─────────────────────────────────────────────
#  DDL — esquema de tablas
# ─────────────────────────────────────────────
SQL_CREATE_RUNS = """
CREATE TABLE IF NOT EXISTS scraping_runs (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    inicio               TEXT NOT NULL,
    fin                  TEXT,
    anno                 INTEGER,
    paginas_extraidas    INTEGER DEFAULT 0,
    registros_nuevos     INTEGER DEFAULT 0,
    registros_ignorados  INTEGER DEFAULT 0,
    registros_totales_db INTEGER DEFAULT 0,
    estado               TEXT DEFAULT 'EN_PROCESO'   -- OK | ERROR | INTERRUMPIDO
);
"""

SQL_CREATE_PROC = """
CREATE TABLE IF NOT EXISTS procedimientos (
    id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    -- Clave de deduplicacion compuesta
    nomenclatura            TEXT    NOT NULL,
    fecha_publicacion       TEXT    NOT NULL,
    -- Datos principales
    entidad                 TEXT,
    reiniciado_desde        TEXT,
    objeto_contratacion     TEXT,
    descripcion_objeto      TEXT,
    codigo_snip             TEXT,
    codigo_unico_inversion  TEXT,
    avance                  TEXT,
    version_seace           TEXT,
    anno_convocatoria       INTEGER,
    -- Campos extra (columnas adicionales capturadas dinamicamente)
    datos_extra             TEXT,   -- JSON con columnas no mapeadas
    -- Auditoria
    primera_vez_visto       TEXT    NOT NULL,
    ultima_actualizacion    TEXT    NOT NULL,
    run_id                  INTEGER REFERENCES scraping_runs(id),

    -- Restriccion de unicidad: misma nomenclatura + misma fecha = duplicado real
    UNIQUE (nomenclatura, fecha_publicacion)
);
"""

SQL_CREATE_IDX = [
    "CREATE INDEX IF NOT EXISTS idx_proc_nomenclatura  ON procedimientos(nomenclatura);",
    "CREATE INDEX IF NOT EXISTS idx_proc_anno          ON procedimientos(anno_convocatoria);",
    "CREATE INDEX IF NOT EXISTS idx_proc_entidad       ON procedimientos(entidad);",
    "CREATE INDEX IF NOT EXISTS idx_proc_objeto        ON procedimientos(objeto_contratacion);",
    "CREATE INDEX IF NOT EXISTS idx_proc_fecha         ON procedimientos(fecha_publicacion);",
]


# ─────────────────────────────────────────────
#  Mapeo dinamico de columnas del portal → DB
# ─────────────────────────────────────────────
def _mapear_columnas(row: dict, anno: int, run_id: int) -> dict:
    """
    Recibe el dict crudo del scraper y devuelve el dict listo para INSERT.
    Columnas no reconocidas van al campo JSON `datos_extra`.
    """
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Busqueda flexible de columnas por palabras clave
    def buscar(row, *palabras):
        palabras = [p.lower() for p in palabras]
        for k, v in row.items():
            kl = k.lower()
            if all(p in kl for p in palabras):
                return (v or "").strip()
        return ""

    nomenclatura   = buscar(row, "nomenclatura") or buscar(row, "nomencla")
    fecha_pub      = buscar(row, "fecha") or buscar(row, "publicacion")
    entidad        = buscar(row, "entidad") or buscar(row, "sigla")
    reiniciado     = buscar(row, "reiniciado")
    objeto         = buscar(row, "objeto", "contrat") or buscar(row, "objeto")
    descripcion    = buscar(row, "descripcion") or buscar(row, "objeto")
    snip           = row.get("Codigo SNIP", "") or buscar(row, "snip")
    inversion      = row.get("Codigo Unico Inversion", "") or buscar(row, "inversion")
    avance         = buscar(row, "avance")
    version        = buscar(row, "version")

    # Columnas conocidas para separar de datos_extra
    claves_mapeadas = {
        nomenclatura, fecha_pub, entidad, reiniciado,
        objeto, descripcion, snip, inversion, avance, version,
        "Codigo SNIP", "Codigo Unico Inversion", "N°", "N", "#",
    }
    datos_extra = {
        k: v for k, v in row.items()
        if k not in claves_mapeadas and v and str(v).strip()
    }

    return {
        "nomenclatura":           nomenclatura,
        "fecha_publicacion":      fecha_pub,
        "entidad":                entidad,
        "reiniciado_desde":       reiniciado,
        "objeto_contratacion":    objeto,
        "descripcion_objeto":     descripcion,
        "codigo_snip":            snip,
        "codigo_unico_inversion": inversion,
        "avance":                 avance,
        "version_seace":          version,
        "anno_convocatoria":      anno,
        "datos_extra":            json.dumps(datos_extra, ensure_ascii=False) if datos_extra else None,
        "primera_vez_visto":      ahora,
        "ultima_actualizacion":   ahora,
        "run_id":                 run_id,
    }


# ─────────────────────────────────────────────
#  API publica
# ─────────────────────────────────────────────

def inicializar_db(ruta: str = DB_PATH) -> sqlite3.Connection:
    """
    Abre (o crea) la base de datos SQLite, aplica el esquema y devuelve
    la conexion con row_factory para acceder por nombre de columna.
    """
    conn = sqlite3.connect(ruta, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")   # mejor escritura concurrente
    conn.execute("PRAGMA foreign_keys=ON;")
    conn.execute(SQL_CREATE_RUNS)
    conn.execute(SQL_CREATE_PROC)
    for idx_sql in SQL_CREATE_IDX:
        conn.execute(idx_sql)
    conn.commit()
    return conn


def iniciar_run(conn: sqlite3.Connection, anno: int) -> int:
    """Registra el inicio de una ejecucion y devuelve su run_id."""
    cur = conn.execute(
        "INSERT INTO scraping_runs (inicio, anno, estado) VALUES (?, ?, 'EN_PROCESO')",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), anno)
    )
    conn.commit()
    return cur.lastrowid


def guardar_batch(conn: sqlite3.Connection, datos: list[dict],
                  run_id: int, anno: int) -> dict:
    """
    Inserta una lista de registros en `procedimientos`.
    Usa INSERT OR IGNORE para respetar la restriccion UNIQUE (nomenclatura, fecha_publicacion).

    Retorna estadisticas: {'nuevos': N, 'ignorados': M}
    """
    nuevos = 0
    ignorados = 0

    SQL_INSERT = """
    INSERT OR IGNORE INTO procedimientos (
        nomenclatura, fecha_publicacion, entidad, reiniciado_desde,
        objeto_contratacion, descripcion_objeto, codigo_snip,
        codigo_unico_inversion, avance, version_seace, anno_convocatoria,
        datos_extra, primera_vez_visto, ultima_actualizacion, run_id
    ) VALUES (
        :nomenclatura, :fecha_publicacion, :entidad, :reiniciado_desde,
        :objeto_contratacion, :descripcion_objeto, :codigo_snip,
        :codigo_unico_inversion, :avance, :version_seace, :anno_convocatoria,
        :datos_extra, :primera_vez_visto, :ultima_actualizacion, :run_id
    )
    """

    for raw in datos:
        if not raw:
            continue
        fila = _mapear_columnas(raw, anno, run_id)

        # Validar que tengamos al menos nomenclatura o fecha para identificar el registro
        if not fila["nomenclatura"] and not fila["fecha_publicacion"]:
            ignorados += 1
            continue

        cur = conn.execute(SQL_INSERT, fila)
        if cur.rowcount > 0:
            nuevos += 1
        else:
            ignorados += 1  # duplicado real (misma nomenclatura + misma fecha)

    conn.commit()
    return {"nuevos": nuevos, "ignorados": ignorados}


def finalizar_run(conn: sqlite3.Connection, run_id: int,
                  paginas: int, stats: dict, estado: str = "OK"):
    """Actualiza los metadatos de la ejecucion al terminar."""
    total_db = conn.execute("SELECT COUNT(*) FROM procedimientos").fetchone()[0]
    conn.execute("""
        UPDATE scraping_runs SET
            fin                  = ?,
            paginas_extraidas    = ?,
            registros_nuevos     = ?,
            registros_ignorados  = ?,
            registros_totales_db = ?,
            estado               = ?
        WHERE id = ?
    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        paginas,
        stats.get("nuevos", 0),
        stats.get("ignorados", 0),
        total_db,
        estado,
        run_id,
    ))
    conn.commit()
    return total_db


def consultar(conn: sqlite3.Connection, anno: int = None,
              entidad: str = None, objeto: str = None,
              texto: str = None, limit: int = 5000) -> list:
    """
    Query flexible. Todos los filtros son opcionales.
    Retorna lista de sqlite3.Row (accesible como dict).
    """
    condiciones = []
    params = []

    if anno:
        condiciones.append("anno_convocatoria = ?")
        params.append(anno)
    if entidad:
        condiciones.append("entidad LIKE ?")
        params.append(f"%{entidad}%")
    if objeto:
        condiciones.append("objeto_contratacion LIKE ?")
        params.append(f"%{objeto}%")
    if texto:
        condiciones.append(
            "(nomenclatura LIKE ? OR descripcion_objeto LIKE ? OR entidad LIKE ?)"
        )
        params.extend([f"%{texto}%"] * 3)

    where = ("WHERE " + " AND ".join(condiciones)) if condiciones else ""
    sql = f"""
        SELECT * FROM procedimientos
        {where}
        ORDER BY fecha_publicacion DESC, id DESC
        LIMIT ?
    """
    params.append(limit)
    return conn.execute(sql, params).fetchall()


def resumen_db(conn: sqlite3.Connection) -> dict:
    """Devuelve estadisticas globales de la base de datos."""
    total     = conn.execute("SELECT COUNT(*) FROM procedimientos").fetchone()[0]
    por_anno  = conn.execute(
        "SELECT anno_convocatoria, COUNT(*) as n FROM procedimientos "
        "GROUP BY anno_convocatoria ORDER BY anno_convocatoria DESC"
    ).fetchall()
    por_tipo  = conn.execute(
        "SELECT objeto_contratacion, COUNT(*) as n FROM procedimientos "
        "WHERE objeto_contratacion != '' GROUP BY objeto_contratacion ORDER BY n DESC"
    ).fetchall()
    runs      = conn.execute(
        "SELECT id, inicio, fin, registros_nuevos, registros_ignorados, estado "
        "FROM scraping_runs ORDER BY id DESC LIMIT 10"
    ).fetchall()
    return {
        "total":    total,
        "por_anno": [dict(r) for r in por_anno],
        "por_tipo": [dict(r) for r in por_tipo],
        "runs":     [dict(r) for r in runs],
    }


# ─────────────────────────────────────────────
#  Exportaciones
# ─────────────────────────────────────────────

def exportar_csv(conn: sqlite3.Connection, anno: int = None,
                 ruta: str = None) -> str:
    """Exporta a CSV los registros filtrados por año (o todos)."""
    filas = consultar(conn, anno=anno, limit=999999)
    if not filas:
        return None

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sufijo = f"_{anno}" if anno else "_todos"
    ruta = ruta or os.path.join(EXPORT_DIR, f"SEACE{sufijo}_{ts}.csv")

    cols = filas[0].keys()
    with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows([dict(r) for r in filas])

    return ruta


def exportar_excel(conn: sqlite3.Connection, anno: int = None,
                   ruta: str = None) -> str:
    """Exporta a Excel con formato de cabecera. Requiere pandas+openpyxl."""
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    filas = consultar(conn, anno=anno, limit=999999)
    if not filas:
        return None

    df = pd.DataFrame([dict(r) for r in filas])
    # Eliminar columna datos_extra del Excel (es JSON interno)
    if "datos_extra" in df.columns:
        df = df.drop(columns=["datos_extra"])

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sufijo = f"_{anno}" if anno else "_todos"
    ruta = ruta or os.path.join(EXPORT_DIR, f"SEACE{sufijo}_{ts}.xlsx")

    with pd.ExcelWriter(ruta, engine="openpyxl") as ew:
        sheet = f"SEACE_{anno}" if anno else "SEACE_todos"
        df.to_excel(ew, index=False, sheet_name=sheet)
        ws = ew.sheets[sheet]
        fill = PatternFill("solid", fgColor="1F4E79")
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            max_l = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 10
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_l + 3, 60)

    return ruta


def log_scraping(mensaje: str, nivel: str = "INFO"):
    """Escribe una linea al archivo de log de scraping."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_path = os.path.join(LOG_DIR, f"scraping_{datetime.now().strftime('%Y%m')}.log")
    icons = {"INFO": "[·]", "OK": "[OK]", "ERR": "[!!]", "WARN": "[!]"}
    linea = f"{ts} {icons.get(nivel, '[ ]')} {mensaje}\n"
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(linea)


# ─────────────────────────────────────────────
#  CLI de utilidad (ejecutar directamente)
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  ═══════════════════════════════════════")
    print("    SEACE DB — Resumen de base de datos")
    print("  ═══════════════════════════════════════\n")

    conn = inicializar_db()
    res  = resumen_db(conn)

    print(f"  Total de registros : {res['total']:,}")
    print()
    print("  Por año:")
    for r in res["por_anno"]:
        print(f"    {r['anno_convocatoria']} → {r['n']:,} procedimientos")
    print()
    print("  Por tipo de objeto:")
    for r in res["por_tipo"][:6]:
        print(f"    {r['objeto_contratacion']:<30} {r['n']:,}")
    print()
    print("  Últimas 5 ejecuciones:")
    for r in res["runs"][:5]:
        print(f"    Run #{r['id']:>3} | {r['inicio']} | "
              f"+{r['registros_nuevos']} nuevos | "
              f"{r['registros_ignorados']} duplicados | {r['estado']}")
    print()
    conn.close()
